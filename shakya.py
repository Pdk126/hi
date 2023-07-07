import os
from tkinter import *
from openpyxl import load_workbook
from alright import WhatsApp
import time
import tkinter.scrolledtext as scrolledtext
from tkinter import messagebox

global x




def load_excel_file(text_box):
    if text_box.get("1.0", END)==" "  :
        path = rf"C:\Users\Lenovo\Desktop\{text_box.get('1.0', END).strip()}.xlsx"
        new_path = path.replace("\n", "")
    else :
        path = rf"C:\Users\Lenovo\Desktop\N.xlsx"
        new_path = path.replace("\n", "")




    try:
        global x
        x= new_path
        wb = load_workbook(new_path, data_only=True)
        ws = wb.active

        column_a = ws["A"]
        no_of_rows = 0

        for cell in column_a:
            no_of_rows += 1
        print(no_of_rows)

        max_columns = ws.max_column

        # Print the result
        print(f"The maximum number of columns in the sheet is: {max_columns}")

        for row in range(1, no_of_rows+1):
            D_value = ws[f'D{row}'].value
            F_value = ws[f'F{row}'].value
            E_value = ws[f'E{row}'].value
            C_value = ws[f'C{row}'].value
            B_value = ws[f'B{row}'].value
            sentence = f'{D_value}  {F_value} ACCOUNT NO : {E_value} OF AMOUNT {C_value} ON {B_value}'
            ws[f'G{row}'].value =sentence

            wb.save(filename=path)
        g_items = []
        column_g = ws["G"]
        for cell in column_g:
            g_items.append(cell.value)


        root = Tk()
        root.geometry("500x500")
        root.title("Messages")
        root.config(bg="light blue")

        preview = scrolledtext.ScrolledText(root, undo=True)
        preview['font'] = ('consolas', '12')
        preview.pack(padx=5, pady=5,expand=True, fill='both')
        for item in g_items:
            preview.insert(END, item + "\n"+"\n"+"\n")

        ok = Button(root,width=5,text="OK",command=lambda:[root.destroy()])
        ok.pack(pady=5,padx=5)
        root.mainloop()
    except:
        pass

















def send_whatsapp() :
    try:
        global x
        messenger = WhatsApp()

        wb = load_workbook(x, data_only=True)
        ws = wb.active

        column_a = ws["A"]
        column_g = ws["G"]

        phone_no_list = []
        message_list = []

        for cell in column_a:
            phone_no_list.append(cell.value)

        for cell in column_g:
            message_list.append(cell.value)

        for i in range(len(phone_no_list)):
            try:
                messenger.find_user(phone_no_list[i])
                messenger.send_message(message_list[i])
                time.sleep(4)
            except Exception as e:
                error_message = f"Error sending message to phone number {phone_no_list[i]}: {str(e)}"
                messagebox.showerror("Error", error_message)

        window.destroy()

    except Exception as e:
        error_message = f"An error occurred: {str(e)}"
        messagebox.showerror("Error", error_message)










window = Tk()
window.minsize(300, 300)
window.title("Excel TO Whatsapp")
window.eval('tk::PlaceWindow . center')
window.config(bg="light blue")

excel_label = Label(window, text="Enter excel file name :", bg="light blue", font=(15))
excel_label.pack(pady=20, padx=20)

text_box = Text(window, height=1, width=15)
text_box.pack(padx=20, pady=5)
text_box.focus()

load_button = Button(window, text=" NEFT LOAD EXCEL FILE", command=lambda :[load_excel_file(text_box)])
load_button.pack(pady=10)







send_button = Button(window,text="SEND "
                                 "WHATSAPP",command=send_whatsapp)
send_button.pack(pady=10)


window.mainloop()
